#!/usr/bin/env python
# coding: utf-8

import time
import openpyxl
from openpyxl import Workbook 
from datetime import datetime, timedelta
import numpy

# Category classification by package name
def findctg (app,pkg):
    try : rst = pkgs[pkg.lower().replace(' ','').replace('\t','')]
    except : rst = '카테고리 미분류'
    return rst

# Excluding duplicate elements in the list
def delovr(list1):
    list2 = list(set(list1))
    list2.sort() 
    return list2

# Entering a number of timestamp returns the datetime object containing year, month, day, hour, and minute information
def che_time(sectime, num=1):
    lotime = time.localtime(sectime)
    y,m1,d,h,m2,s, w = \
    int(time.strftime('%Y', lotime)),\
    int(time.strftime('%m', lotime)),\
    int(time.strftime('%d', lotime)),\
    int(time.strftime('%H', lotime)),\
    int(time.strftime('%M', lotime)),\
    int(time.strftime('%S', lotime)),\
    time.strftime('%a', time.localtime(sectime))

    time1 = datetime(y,m1,d,h,m2,s)
    time2 = (time1 + timedelta(hours=9)) 
    t = time2 if num == 2 else time1

    yy,mm1,dd,hh,mm2,ss = t.year, t.month, t.day, t.hour, t.minute, t.second
    midn_time= (datetime(y,m1,d,23,59,59)-datetime(1970,1,1)).total_seconds()+1
    return t, midn_time

# Entering year, month, day, hour, and minute changes to a number in timestamp format
def oclock(y,m,d,h):
    rst = (datetime(y,m,d,h,59,59)-datetime(1970,1,1)).total_seconds()+1
    return rst-(3600*9)

# A function that returns the summed value for the app usage history from 0 to 24 hours by hour
def hourwiselist(mmr, ipt):
    ctg = findctg(ipt[1], ipt[2])

    d0 = che_time(mmr[0])[1]
    d1 = che_time(mmr[0])[0]
    d2 = che_time(ipt[0])[0]
    
    pkgnm = ipt[2]

    if d1.day == d2.day: 
        if d1.hour == d2.hour:
            usetime = (float(ipt[0])-float(mmr[0]))/60
            wkday = d1.weekday()
            rst = [[ctg,mmr[1],d1.hour,usetime,pkgnm,mmr[0],ipt[0]]]
        else:
            rst = []
            h_delta = d2.hour - d1.hour
            for i in range(h_delta +1):
                if i == 0 : 
                    ock = oclock(d1.year, d1.month, d1.day, d1.hour)
                    usesec = (ock - mmr[0])
                    usemin = usesec/60
                elif i == h_delta :
                    usemin = ((ipt[0]-mmr[0]-usesec)%3600)/60
                else:
                    usemin = 60
                wkday = d1.weekday()
                rst.append([ctg, mmr[1],d1.hour+i,usemin,pkgnm,mmr[0],ipt[0]])

    else: 
        rst = []
        d_delta = d2.day - d1.day
        for j in range(d_delta+1):
            if j == 0:
                h_delta = 23 - d1.hour
                for i in range(h_delta +1):
                    if i == 0 : 
                        ock = oclock(d1.year, d1.month, d1.day, d1.hour)
                        usesec = (ock - mmr[0])
                        usemin = usesec/60
                    else:
                        usemin = 60
                    wkday = (d1+timedelta(days=j)).weekday()
                    rst.append([ctg, mmr[1],d1.hour+i,usemin,pkgnm,mmr[0],ipt[0]])
            elif j== d_delta:
                h_delta = d2.hour
                for i in range(h_delta +1):
                    if i == h_delta : 
                        usemin = ((ipt[0]-mmr[0]-usesec )%3600)/60
                    else:
                        usemin = 60
                    wkday = (d1+timedelta(days=j)).weekday()
                    rst.append([ctg, mmr[1],i,usemin,pkgnm,mmr[0],ipt[0]])
            else:
                for i in range(24):
                    usemin = 60
                    wkday = (d1+timedelta(days=j)).weekday()
                    rst.append([ctg, mmr[1],i,usemin,pkgnm,mmr[0],ipt[0]])
    return rst    

# Inserting a list returns sum, average, maximum, minimum, 
# and median values based on the elements of the list. The sum is used when calculating the percentage option.
def statistic(list1):
    list_usetime = [_ for _ in list1 if _!=0 ]
    if list_usetime == []:
        rst = [] 
    else:
        sumv = sum(list_usetime)
        avrg = numpy.mean(list_usetime)
        std = numpy.std(list_usetime)
        med = numpy.median(list_usetime)
        minv = min(list_usetime)
        maxv = max(list_usetime)
        rst = [list1, sumv, avrg, med,maxv,minv]
    return rst 

# It is used to create a group from the baseline data and calculate the weighted average for each group
def sumproduct(aa) :
    usetime , cnt =  0,0
    for a in aa:
        if a ==0 :continue
        usetime += a[0]
        cnt += a[1]
    if usetime == 0: rst = 0
    else:  rst = usetime/cnt
    return rst

# Entering an 8-digit string as a string returns 'day of the week'
def che_wday(date):
    weekdayeng= ['Mon', 'Tue', 'Wed', 'Thu', 'Fri','Sat','Sun']
    weekdaykor= ['월', '화', '수', '목', '금','토','일']
    idx = datetime(int(date[:4]), int(date[5:6]),int(date[6:])).weekday()
    return (idx, weekdayeng[idx], weekdaykor[idx])

#s_week0 ~ s_week6 variables are created using standard data for each day of the week
def make_s_week(standard_week0  ):
    s_week0 = {}
    for ctg_name in grp21.keys(): #---- None값은 0으로 처리
        if ctg_name not in s_week0.keys() : s_week0[ctg_name] = {}
        for cahosucnt in standard_week0 :
            hour = cahosucnt[1][:-1]
            if ctg_name == cahosucnt[0]: 
                if (cahosucnt[2],cahosucnt[3]) == (None, None):
                    add = [0,0]
                else :  add = [cahosucnt[2],cahosucnt[3]]
                s_week0[ctg_name][hour]  = add
    return  s_week0 

# A function that selects the package name of input data that is not in the category list and classifies it as unclassified
def miscategory(inputlist):
    dic01 = makedic01(inputlist)
    missing = []
    for date in dic01.keys():
        for emt4 in dic01[date]:
            if '무시' in findctg(emt4[1], emt4[2]): continue
            if '미분류' in findctg(emt4[1], emt4[2]):
                add = ['missing' , emt4[1], emt4[2]]
                if add not in missing: missing.append(add)
    return missing + ctg_nms

# A function that checks how many years, months, days, and times of the input data the app usage time was recorded.
def finding(inputlist):## 구 input 함수
    dic05 =  makedic05(inputlist)
    
    fninput = [['DATE']+[f'{_}시' for _ in range(24)]]
    for date in dic05.keys():
        hour24 = [0 for _ in range(24)]
        for cat_ry in dic05[date]:
            usetimelist24 =  cat_ry[2:]
    #         print(date, cat_ry[2:])
            for idx in range(24):
                hour24[idx] += usetimelist24[idx]
        fninput.append([date]+hour24)
    #         hour24[idx] += 1
    #     break
    return fninput   

# A function that compares the application time corresponding to the social and communication categories of input data with the reference value data of 80 people
def social(inputlist, timeunit ='h', timeunit_n=1, divisionmethod ='section', comparisonvaluetype='timeofday', comparisionstatistics='mean' ):
    grp70 = digital(inputlist,timeunit, timeunit_n,divisionmethod, comparisonvaluetype,comparisionstatistics )
    grp80 = {}
    sumstd , sumipd = 0,0
    for comsoc in ['Communication', 'Social']:
        if comsoc  in grp70.keys():
            std = grp70[comsoc]['standard']
            ipd = grp70[comsoc ]['input']
            rate = grp70[comsoc ]['compare']
        else: std, ipd,  rate= 0,0,0 

        grp80[comsoc] = {'standard':std , 'input': ipd, 'compare': rate}
        sumstd += std; sumipd += ipd
    grp80['Total'] =    {'standard':sumstd , 'input': sumipd, 'compare': round(sumipd/sumstd, 2)}
    return grp80 

###---------------------------변수 프린팅 용 함수 -------------------------------------

def list2xl(inputlist, filename=''):
    now = time.localtime()
    save_time = "%04d-%02d-%02d_%02d%02d%02d"%(now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    xlfnm =f'output{filename}_{save_time}.xlsx'
    wb = Workbook()
    ws= wb.active
    for ipt in inputlist:
        ws.append(ipt)
    wb.save(filename = xlfnm) 
    print(f'{xlfnm} 파일 저장완료')
    
def dic2xl(aa, filename=''):
    rows = [''] +list(aa[list(aa.keys())[0]].keys())
    rst = []
    for rx, row in enumerate(rows):
        add = [row]
        if rx == 0 : rst.append(add + list(aa.keys()))
        else:
            add = add + [str(aa[a][row]) for a in list(aa.keys())]
            rst.append(add)
    list2xl(rst,filename)     

def var2xl(dic01, filename=''):  
    rst = [] 
    for knm in  dic01.keys():
        for emt4 in dic01[knm]:
            rst.append([knm]+emt4)
    list2xl(rst,filename)
    
def grp2xl(grpN, filename='') :
    list1 = []
    for knm in grpN.keys():
        add = [knm ] #+ grpN[knm]
        for bundle in grpN[knm]:
            add.append( str(bundle))
        list1.append(add)
    list2xl(list1, filename)

def ipt2xl(ipt01, filename=''):
    for i in [1]:
        list1 = []
        for knm in ipt01.keys():
    #         if knm not in [_[0] for _ in list1]: list1.append([[knm]])
            for datetime in ipt01[knm].keys():
                add = [knm, datetime, ipt01[knm][datetime ]]
                list1.append(add)
    list2xl(list1, filename)

def printd5(dic05, filename):  
    rst = [] 
    for knm in  dic05.keys():
        for emt4 in dic05[knm]:
            rst.append([knm]+emt4)
    inputlist = rst
    xlfnm =f'dic05_{filename}.xlsx'
    wb = Workbook()
    ws= wb.active
    for ipt in inputlist:
        ws.append(ipt)
    wb.save(filename = xlfnm) 
    print(f'{xlfnm} 파일 저장완료')

##---------------------------- dic01 / doc04 추출함수 ---------------------------------------

# A function created to extract the variable of dic01 for use in the miscategory function
def makedic01(inputlist):
    dic00 = []
    for ipt in inputlist:
        if ipt[3] not in ['MOVE_TO_FOREGROUND', 'MOVE_TO_BACKGROUND']:  continue 
        if ipt[2] in [_[0] for _ in ignore_pkgapp]: continue ## 무시카테고리의 패키지라면 스킵
        add = [int(ipt[0])/1000, ipt[1], ipt[2], ipt[3]] ## 시간단위 ms > s
        if add not in dic00 : ## 중복 데이터 스킵
            dic00.append(add)      

    dic01 = {} ## 시간단위 ms >> s
    for ptx, inpt in enumerate(dic00 ) :
        dtinf = che_time((inpt[0]))[0]
        date8 = dtinf.year*10000 + dtinf.month*100 + dtinf.day 
        keynm = date8 
        if date8 not in list(dic01.keys()):
            dic01[keynm] = []
            dic01[keynm].append(inpt)
        else: 
            dic01[keynm].append(inpt )
    return  dic01 

# Function created to extract dic01 variable separately for use in finding function
def makedic05(inputlist):
    dic01 = makedic01(inputlist)
    ## 카테고리 분류 / 사용시간 계산   
    dic02 = {} ## 시간단위  second >> minute
    for d1 in list(dic01.keys()):
        input4list  = dic01[d1]

        listed1, mmr = [],[]
        for ipt in input4list: 
            try : tp = type(mmr[0]) 
            except : tp = 'none'

            if tp == float or tp == 'none':
                if mmr == [] : 
                    if ipt[3] == stimechk : mmr = ipt

                else:
                    if mmr[1] == ipt [1]:
                        if ipt[3] == etimechk: 
                            listed1 += hourwiselist(mmr, ipt)
                            mmr = []
                        else: pass
                    else:
                        if ipt[3] == stimechk: mmr.append(ipt)
                        else: pass
            else:
                if int[1] in [_[1] for _ in mmr]:
                    mx = [_[1] for _ in mmr].index(int[1])
                    if ipt[3] == stimechk:
                        del mmr[mx]
                        mmr.append(ipt)
                    else:
                        listed1 += hourwiselist(mmr[mx], ipt)
                else: 
                    if ipt[3] == stimechk: mmr.append(ipt)
                    else: pass
        dic02[d1] = listed1

    ## [카테고리명, 어플명, 24개 사용시간]
    dic03 = {}
    ex_lt1 = []
    for d2 in list(dic02.keys()):
        if d2 not in dic03.keys(): dic03[d2] = []
        for emt in dic02[d2]:
            hour24ap = [0 for _ in range(24)]
            hour = emt[2]
            hour24ap[hour] += emt[3]
            add = [emt[0], f'{emt[1]}-{emt[4]}']+hour24ap 
            dic03[d2].append(add) 

    ## [카데고리 , total, 24개 사용시간]
    dic04 = {}
    for d3 in list(dic03.keys()):
        listed2 = dic03[d3]
        listed3 = []
        for ctnm in allctgnms : #delovr([_[0] for _ in ctg_nms]):
            hour24ct = [0 for _ in range(24)]
            for lt2 in listed2:
                if ctnm == lt2[0]:
                    for xx, ea in enumerate(lt2[2:]):
                        hour24ct[xx] += ea
            listed3.append([ctnm,'Total']+hour24ct)
        dic04[d3] = listed3
    ## 날짜 순서대로 정렬
    days_inputdata = list(dic04.keys())
    days_inputdata.sort()
    dic05 = {}
    for dayipt in days_inputdata:
        dic05[dayipt] = dic04[dayipt]
    return dic05

# A function to determine the usage time of an app category and compare it with the comparison value for quantification of the user's digital behavior
def digital(inputlist, timeunit ='h', timeunit_n=1, divisionmethod ='section', comparisonvaluetype='timeofday', comparisionstatistics='mean' ):
    global grp21
    if  timeunit not in ['h', 'am', 'pm', 'all', 'd'] or \
    type(timeunit_n) != int or divisionmethod not in ['section', 'nonsection'] or\
    comparisonvaluetype not in ['timeofday', 'dayofweek'] or \
    comparisionstatistics not in ['mean', 'median', 'max', 'min', 'percentage', \
                                  '_mean', '_median', '_max', '_min', '_percentage']:
        raise Exception('유효하지 않은 파라미터 입력')

    tu, tun = timeunit , timeunit_n
    if divisionmethod == 'section': absrel = 'abs'
    else: absrel = 'rel'
    if comparisonvaluetype==  'timeofday': hourweek = 'hour'
    else :  hourweek ='week'
    comparisionstatistics00 = comparisionstatistics
    comparisionstatistics = comparisionstatistics.replace('_','')
    if comparisionstatistics in [ 'mean', 'max','min']:
        sttstc =comparisionstatistics
    elif comparisionstatistics == 'median' : sttstc ='med'
    else: sttstc ='per'

    ## **************새 양식의 파일의 경우 Activity_paused, Activity_resumed로 변경 *******************       
    stimechk, etimechk ='MOVE_TO_FOREGROUND', 'MOVE_TO_BACKGROUND'  

    ## { 날짜: (datetime, [ input lslt] ) }
    ## stimechk, etimechk 데이터만 남기기 
    dic00 = []
    for ipt in inputlist:
        if ipt[3] not in ['MOVE_TO_FOREGROUND', 'MOVE_TO_BACKGROUND']:  continue 
        if ipt[2] in [_[0] for _ in ignore_pkgapp]: continue ## 무시카테고리의 패키지라면 스킵
        add = [int(ipt[0])/1000, ipt[1], ipt[2], ipt[3]] ## 시간단위 ms > s
        if add not in dic00 : ## 중복 데이터 스킵
            dic00.append(add)      

    dic01 = {} ## 시간단위 ms >> s
    for ptx, inpt in enumerate(dic00 ) :
        dtinf = che_time((inpt[0]))[0]
        date8 = dtinf.year*10000 + dtinf.month*100 + dtinf.day 
        keynm = date8 
        if date8 not in list(dic01.keys()):
            dic01[keynm] = []
            dic01[keynm].append(inpt)
        else: 
            dic01[keynm].append(inpt )

    ## 카테고리 분류 / 사용시간 계산   
    dic02 = {} ## 시간단위  second >> minute
    for d1 in list(dic01.keys()):
        input4list  = dic01[d1]

        listed1, mmr = [],[]
        for ipt in input4list: 
            try : tp = type(mmr[0]) 
            except : tp = 'none'

            if tp == float or tp == 'none':
                if mmr == [] : 
                    if ipt[3] == stimechk : mmr = ipt

                else:
                    if mmr[1] == ipt [1]:
                        if ipt[3] == etimechk: 
                            listed1 += hourwiselist(mmr, ipt)
                            mmr = []
                        else: pass
                    else:
                        if ipt[3] == stimechk: mmr.append(ipt)
                        else: pass
            else:
                if int[1] in [_[1] for _ in mmr]:
                    mx = [_[1] for _ in mmr].index(int[1])
                    if ipt[3] == stimechk:
                        del mmr[mx]
                        mmr.append(ipt)
                    else:
                        listed1 += hourwiselist(mmr[mx], ipt)
                else: 
                    if ipt[3] == stimechk: mmr.append(ipt)
                    else: pass
        dic02[d1] = listed1

    ## [카테고리명, 어플명, 24개 사용시간]
    dic03 = {}
    ex_lt1 = []
    for d2 in list(dic02.keys()):
        if d2 not in dic03.keys(): dic03[d2] = []
        for emt in dic02[d2]:
            hour24ap = [0 for _ in range(24)]
            hour = emt[2]
            hour24ap[hour] += emt[3]
            add = [emt[0], f'{emt[1]}-{emt[4]}']+hour24ap 
            dic03[d2].append(add)

    ## [카데고리 , total, 24개 사용시간]
    dic04 = {}
    for d3 in list(dic03.keys()):
        listed2 = dic03[d3]
        listed3 = []
        for ctnm in allctgnms : #delovr([_[0] for _ in ctg_nms]):
            hour24ct = [0 for _ in range(24)]
            for lt2 in listed2:
                if ctnm == lt2[0]:
                    for xx, ea in enumerate(lt2[2:]):
                        hour24ct[xx] += ea
            listed3.append([ctnm,'Total']+hour24ct)
        dic04[d3] = listed3

    ## 날짜 순서대로 정렬
    days_inputdata = list(dic04.keys())
    days_inputdata.sort()
    dic05 = {}
    for dayipt in days_inputdata:
        dic05[dayipt] = dic04[dayipt]

    ## {카테고리명 : {20190509 : (datetiem, [24시간리스트])}}로 변경
    dic06 = {}
    for ctg_name in delovr([_[0] for _ in ctg_nms]):
        mmrdic06 = {}
        for mday in list(dic05.keys()):
            mmrdic06list = []
            for d05 in dic05[mday]:
                if ctg_name == d05[0] : mmrdic06list = d05[2:]
            mmrdic06[mday]  = mmrdic06list
        dic06[ctg_name] = mmrdic06

    ### ----------------------- 인풋데이터 처리 ---------------------------------
    ipt01 = {}
    for ctg_name in allctgnms:
        for date in dic06[ctg_name].keys():
            if ctg_name not in ipt01.keys() :
                ipt01[ctg_name] = {}
            for i in range(24):
                if absrel == 'rel': 
                    if dic06[ctg_name][date][i] != 0:
                        ipt01[ctg_name][f'{date}h{i}'] = dic06[ctg_name][date][i] 
                else: # absrel == 'abs'
                    ipt01[ctg_name][f'{date}h{i}'] = dic06[ctg_name][date][i] 

    ## 묶음만들기 
    grp10 = {}
    if tu in ['h','d','all']:
        for ctg_name in allctgnms:
            if tu == 'd' : tun = tun*24
            elif tu == 'all': tun = len(ipt01[ctg_name].keys())

            if ctg_name not in grp10.keys(): grp10[ctg_name] = []
            for dhx, datehour in enumerate(ipt01[ctg_name].keys()):
                add = [datehour, ipt01[ctg_name][datehour]]
                if tun ==1 :
                    mmr =[]
                    mmr.append(add )
                    grp10[ctg_name].append(mmr)
                else:
                    if dhx % tun == 0:
                        mmr =[]
    #                     mmr.append(add )
    #                 elif (dhx+1)% tun == 0:

                    if (dhx+1)% tun == 0:
                        mmr.append(add )
                        grp10[ctg_name].append(mmr)
                    else: 
                        if dhx+1 == len(ipt01[ctg_name].keys()) :
                            mmr.append(add )
                            grp10[ctg_name].append(mmr)
                        else: mmr.append(add )
    elif tu in ['am', 'pm']:
        for ctg_name in allctgnms:
            if ctg_name not in grp10.keys(): grp10[ctg_name] = []

            for dhx, datehour in enumerate (ipt01[ctg_name].keys()):
                date = int(datehour.split('h')[0])
                hour = int(datehour.split('h')[1])
                if tu == 'am' and hour >=12: continue
                if tu == 'pm' and hour <12 : continue
                ex_dates = [_[0][0].split('h')[0] for _ in grp10[ctg_name]]
                if str(date) not in ex_dates :
                    grp10[ctg_name].append([])
                grp10[ctg_name][-1].append([datehour ,ipt01[ctg_name][datehour]])
    grp11 = grp10            

    # 각 묶음의 합계를 원소로 하는 집합 만들기
    grp20 = {}
    for ctg_name in allctgnms:
        if ctg_name not in grp20.keys(): grp20[ctg_name] = []
        for bundle in grp11[ctg_name]:
            grp20[ctg_name].append (sum([_[1]for _ in bundle]))

    ##기준데이터에서 값 추출 시 필요한 정보 만들기        
    grp21 = {}
    for ctg_name in allctgnms:
        for bundle in grp11[ctg_name]:
            if ctg_name not in grp21.keys(): grp21[ctg_name] = []
            add = [_[0]for _ in bundle if _[1] != 0 ] 
            if add  != []:  grp21[ctg_name].append(add) 
    for ctg_name in allctgnms:
        try:  
            if grp21[ctg_name] == [] : del grp21[ctg_name]
        except: pass


    # 집합에서 통계값 추출 / 통계값 없는 카테고리 삭제
    grp30 = {}
    sumforper = 0
    for ctg_name in allctgnms:
        rst =  statistic(grp20[ctg_name])
        if rst != []: 
            if ctg_name not in grp30.keys(): grp30[ctg_name] = {}
            if sttstc == 'mean': grp30[ctg_name]['result'] = rst[2]
            elif sttstc == 'med': grp30[ctg_name]['result'] = rst[3]
            elif sttstc == 'max': grp30[ctg_name]['result'] = rst[4]
            elif sttstc == 'min': grp30[ctg_name]['result'] = rst[5]
            else: # sttstc = 'per': 
                grp30[ctg_name]['result'] = rst[1]  ## 합계로 퍼센테이지계산
    #             grp30[ctg_name]['result'] = rst[2]  ## 평균으로 퍼센테이지계산
            sumforper += grp30[ctg_name]['result']
    if sttstc == 'per': 
        for ctg_name in grp30.keys():
            grp30[ctg_name]['result'] = grp30[ctg_name]['result'] /sumforper*100
        grp30['Total'] = {'result': 100 }    
    else:
        grp30['Total'] = {'result' : sumforper}        


    ### ----------------------- 기준데이터 처리 ---------------------------------
    if hourweek == 'hour':
        s_hour = {}
        for ctg_name in grp21.keys(): #---- None값은 0으로 처리
            if ctg_name not in s_hour.keys() : s_hour[ctg_name] = {}
            for cahosucnt in standard_hour:
                hour = cahosucnt[1][:-1]
                if ctg_name == cahosucnt[0]: 
                    if (cahosucnt[2],cahosucnt[3]) == (None, None):
                        add = [0,0]
                    else :  add = [cahosucnt[2],cahosucnt[3]]
                    s_hour[ctg_name][hour]  = add

        ##  묶음 만들기 
        grp40 = {}
        for ctg_name in grp21.keys():    
            if ctg_name not in grp40.keys(): grp40[ctg_name] =[]
        #     if len(grp40[ctg_name]) == 0 : grp40[ctg_name].append([])
            for datehours in grp21[ctg_name ]:
                grp40[ctg_name].append([])
                for datehour in datehours:
                    strhour =  datehour.split('h')[1]  
                    add =  s_hour[ctg_name][strhour] 
                    grp40[ctg_name][-1].append(add)
    else: ## week인 경우
        standard_week0, standard_week1,standard_week2 , standard_week3, standard_week4, \
        standard_week5, standard_week6 = [['Category', 'Time', 'Sum', 'Count']],\
        [['Category', 'Time', 'Sum', 'Count']],[['Category', 'Time', 'Sum', 'Count']],\
        [['Category', 'Time', 'Sum', 'Count']],[['Category', 'Time', 'Sum', 'Count']],\
        [['Category', 'Time', 'Sum', 'Count']],[['Category', 'Time', 'Sum', 'Count']]
        for wx , emt in enumerate(standard_week):
            if wx in [0,1] : continue
            standard_week0.append([emt[0], emt[1], emt[2], emt[3] ])
            standard_week1.append([emt[0], emt[1], emt[4], emt[5] ])
            standard_week2.append([emt[0], emt[1], emt[6], emt[7] ])
            standard_week3.append([emt[0], emt[1], emt[8], emt[9] ])
            standard_week4.append([emt[0], emt[1], emt[10], emt[11]] )
            standard_week5.append([emt[0], emt[1], emt[12], emt[13]] )
            standard_week6.append([emt[0], emt[1], emt[14], emt[15]] )
        s_week0, s_week1, s_week2, s_week3, s_week4, s_week5, s_week6 = make_s_week(standard_week0),\
        make_s_week(standard_week1), make_s_week(standard_week2), make_s_week(standard_week3),\
        make_s_week(standard_week4), make_s_week(standard_week5), make_s_week(standard_week6)
        s_week_list = [s_week0, s_week1, s_week2, s_week3, s_week4, s_week5, s_week6 ]
        ##  묶음 만들기 
        grp40 = {}
        for ctg_name in grp21.keys():    
            if ctg_name not in grp40.keys(): grp40[ctg_name] =[]
        #     if len(grp40[ctg_name]) == 0 : grp40[ctg_name].append([])
            for datehours in grp21[ctg_name ]:
                grp40[ctg_name].append([])
                for datehour in datehours:
                    wday = che_wday(datehour[:8] )
                    strhour =  datehour.split('h')[1]  
                    add = s_week_list[wday[0]][ctg_name][strhour] 
                    grp40[ctg_name][-1].append(add)

    ## 묶음의 가중평균(sumproduct)를 원소로 하는 집합 만들기 
    grp50 = {}
    for ctg_name in grp40.keys():
        if ctg_name not in grp50.keys(): grp50[ctg_name] =[]
        for bundle in grp40[ctg_name]:
            grp50[ctg_name].append (sumproduct(bundle))

    # 집합에서 통계값 추출 
    grp60 = {}

    sumforper = 0
    for ctg_name in grp50.keys() :
        rst =  statistic(grp50[ctg_name])
        if rst != []: 
            if ctg_name not in grp60.keys(): grp60[ctg_name] = {}
            if sttstc == 'mean': grp60[ctg_name]['result'] = rst[2]
            elif sttstc == 'med': grp60[ctg_name]['result'] = rst[3]
            elif sttstc == 'max': grp60[ctg_name]['result'] = rst[4]
            elif sttstc == 'min': grp60[ctg_name]['result'] = rst[5]
            else: # sttstc = 'per': 
                grp60[ctg_name]['result'] = rst[1]  # 합계로 퍼센테이지 계산
    #             grp60[ctg_name]['result'] = rst[2]   ## 평균으로 퍼센테이지 계산
            sumforper += grp60[ctg_name]['result'] 

    if sttstc == 'per': 
        for ctg_name in grp60.keys():
            grp60[ctg_name]['result'] = grp60[ctg_name]['result'] /sumforper*100
        grp60['Total'] = {'result': 100 }    
    else:
        grp60['Total'] = {'result' : sumforper}        

    ### ----------------------- 아웃풋 테이블 만들기  ---------------------------------
    grp70 = {}
    for ctg_name in grp30.keys():
        if ctg_name not in grp70.keys() : grp70[ctg_name] = {}
        try : std = round(grp60[ctg_name]['result'],2)
        except: std = ''
        ipd = round(grp30[ctg_name]['result'],2)
        if std in [0, ''] : rate = ''
        else :    rate = round(ipd/std,2)
        grp70[ctg_name] = {'standard':std , 'input': ipd, 'compare': rate}

    if '_' not in comparisionstatistics00:
        final = grp70    
    else: 
        final = {'grp70':grp70, 'grp60':grp60, 'grp50':grp50, 'grp40':grp40, 'grp30':grp30, 'grp21':grp21, \
                    'grp20':grp20, 'grp10':grp10, 'ipt01':ipt01, 'dic06':dic06, 'dic05':dic05,'dic04':dic04, \
                    'dic03':dic03, 'dic02':dic02, 'dic01':dic01, 'standard_hour': standard_hour, 'standard_week':standard_week }
    return final



##----------------------------- 추후 내장 테이터로 작업할 영역 ------------------------------ 

from openpyxl import load_workbook

## 기준데이터 가져오기 (시간별)
wb = load_workbook('hcoding\\standarddata_hourwise.xlsx' , data_only=True)
ws = wb.active 
standard_hour = []
for rx, row in enumerate(ws.rows):
    a,b,c,d = row[0].value, row[1].value, row[2].value,row[3].value
    standard_hour.append([a,b,c,d])
wb.close()
    
## 기준데이터 가져오기 (요일별)
wb = load_workbook('hcoding\\standarddata_weekwise.xlsx' , data_only=True)
ws = wb.active 
standard_week = []
for rx, row in enumerate(ws.rows):
    a,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p = \
    row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value,\
    row[6].value, row[7].value, row[8].value, row[9].value, row[10].value, row[11].value,\
    row[12].value, row[13].value, row[14].value, row[15].value
    standard_week.append([a,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p]) 
wb.close()


## 카테고리 리스트  가져오기
wbctg = load_workbook('hcoding\\category.xlsx' , data_only=True)
wsctg = wbctg.active
ctg_nms = []
for rx, row in enumerate(wsctg.rows):
    if rx == 0 or row[0].value == None: continue
    add = [row[0].value, row[1].value, row[2].value]
    ctg_nms.append(add)
    
ctgs, pkgs = {}, {}
for ea in ctg_nms:
    if ea[1] != None: ctgs[ea[1].lower().replace(' ','').replace('\t','')] = ea[0]
    if ea[2] != None: pkgs[ea[2].lower().replace(' ','').replace('\t','')] = ea[0]
allctgnms = delovr([_[0] for _ in ctg_nms  if _[0] not in ['무시']])
ignore_pkgapp = [[_[2],_[1]] for _ in ctg_nms if _[0] == '무시']            
###################################################################################
